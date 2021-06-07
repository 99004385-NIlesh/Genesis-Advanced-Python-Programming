
"""
This is the main file of the project
"""
import openpyxl
import active_excel
from print_psnumber import *
from input_details import *
from sheets import *

class Project:
    """
    This is the Class
    """
    columns = []
    def __init__(self, ps_number, sheet) -> None:
        self.ps_number = ps_number
        self.sheet = sheet

    def columns_list(self):
        """ This function will provide the list of all the columns """
        req_sheet = WORKBOOK[self.sheet]
        for i in range(1, req_sheet.max_column+1):
            column_name = req_sheet.cell(row=1, column=i)
            self.columns.append(column_name.value)
        return self.columns
    def access_sheet(self):
        """ This function will access the particular sheet and gives required output """
        features_list = []
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        req_sheet = active_excel.WORKBOOK[self.sheet]
        for i in range(1, req_sheet.max_column+1):
            feature = new_sheet.cell(row=1, column=i)
            feature.value = self.columns[i-1]
        for i in range(1, req_sheet.max_row+1):
            if req_sheet.cell(row=i, column=1).value == self.ps_number:
                for j in range(1, req_sheet.max_column+1):
                    features_list.append(req_sheet.cell(row=i, column=j).value)
        for i in range(1, len(features_list)+1):
            feature = new_sheet.cell(row=2, column=i)
            feature.value = features_list[i-1]


        new_workbook.save("output.xlsx")
        print()
        print("New Excel File created successfully in the same directory")
if __name__ == "__main__":
    print("Here is the list of all PS Numbers:-\n")
    PS_NUMBER = print_psnumber_list()
    print(PS_NUMBER)
    print()
    print("Here is the list of all the sheets:-\n")
    SHEETS = print_sheets()
    print(SHEETS)
    print()
    REQUIRED_PS_NUMBER, REQUIRED_SHEET_NAME = input_details()

    try:
        if REQUIRED_PS_NUMBER in PS_NUMBER and REQUIRED_SHEET_NAME in SHEETS:
            OBJECT = Project(REQUIRED_PS_NUMBER, REQUIRED_SHEET_NAME)
        else:
            raise Exception
    except:
        print(
            """
            The Details you entered does not exist
            Kindly enter again carefully
            """
        )
    else:
        OBJECT.columns_list()
        OBJECT.access_sheet()
        
